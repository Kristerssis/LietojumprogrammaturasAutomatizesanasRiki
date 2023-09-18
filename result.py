from openpyxl import Workbook, load_workbook 
workBook=load_workbook('tests/test1.xlsx') 
workSheet=workBook.active 

# write your code here
max_row=workSheet.max_row

for i in range(2,max_row):
    hours=workSheet['B'+str(i)].value
    rate=workSheet['C'+str(i)].value
    if (type(hours)!=str and type(rate)!=str):    
        salary=float(hours)*float(rate)
        workSheet['D'+str(i)].value=salary
        #print(salary)
    else:
        salary = 'Nan'
        workSheet['D'+str(i)].value=salary

workBook.save('tests/result.xlsx')
workBook.close()
